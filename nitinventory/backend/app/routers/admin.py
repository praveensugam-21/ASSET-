from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Request
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.limiter import limiter
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from datetime import datetime
from typing import Optional, List
from app.core.config import settings

from app.core.database import get_db
from app.core.deps import require_roles, get_current_user
from app.core.security import get_password_hash
from app.models.user import User, Department, RoleManager
from app.models.budget import BudgetMaster, FinancialYear, PurchaseCategory, ProcurementManager, PhaseManager, Settings
from app.models.purchase_request import WorkFlowHierarchy

router = APIRouter(prefix="/api/admin", tags=["admin"])
AdminDep = Depends(require_roles("admin"))
DeanOrAdminDep = Depends(require_roles("admin", "dean_approver", "apex_approver"))
DeanBudgetDep = Depends(require_roles("dean_approver"))
BudgetViewDep = Depends(require_roles("admin", "dean_approver", "hod", "faculty", "apex_approver"))


# ─────────────────────────────────────────────────────────────────────────────
# USERS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/users")
async def list_users(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).order_by(User.name))
    users = result.scalars().all()
    return [
        {
            "id": u.id,
            "name": u.name,
            "email": u.email,
            "designation": u.designation,
            "gender": u.gender,
            "role_id": u.role_id,
            "department_id": u.department_id,
            "is_active": u.is_active,
            "is_approved": u.is_approved,
            "last_login_at": u.last_login_at.isoformat() if u.last_login_at else None,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        }
        for u in users
    ]


@router.post("/users")
async def create_user(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    existing = await db.execute(select(User).where(User.email == body["email"].lower()))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Email already in use")
    u = User(
        name=body["name"],
        email=body["email"].lower(),
        hashed_password=get_password_hash(body["password"]),
        designation=body.get("designation", ""),
        gender=body.get("gender", "male"),
        role_id=body.get("role_id"),
        department_id=body.get("department_id"),
        is_active=True,
        is_approved=body.get("is_approved", True),
    )
    db.add(u)
    await db.commit()
    return {"message": "User created", "id": u.id}


@router.get("/users/import-template")
async def import_template(_=AdminDep):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Users Import"

    # Set headers
    headers = ["Name", "Email", "Department Name", "Role Name"]
    ws.append(headers)

    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers))):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    # Sample rows
    samples = [
        ["John Doe", "john.doe@nitt.edu", "Computer Science and Engineering", "Faculty"],
        ["Jane Smith", "jane.smith@nitt.edu", "Electronics and Communication Engineering", "Head of Department"],
    ]
    for row in samples:
        ws.append(row)

    # Style samples
    for row_idx in range(2, len(samples) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="users_import_template.xlsx"'}
    )


@router.post("/users/import")
async def import_users(file: UploadFile, db: AsyncSession = Depends(get_db), _=AdminDep):
    import io
    from openpyxl import load_workbook
    from app.models.user import User, Department, RoleManager
    from app.core.security import get_password_hash
    from sqlalchemy import select

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xls) are supported")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The sheet is empty")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    expected = ["name", "email", "department name", "role name"]
    
    # Check that headers match
    for exp in expected:
        if exp not in headers:
            raise HTTPException(status_code=400, detail=f"Missing column '{exp}' in header row")

    name_idx = headers.index("name")
    email_idx = headers.index("email")
    dept_idx = headers.index("department name")
    role_idx = headers.index("role-name") if "role-name" in headers else headers.index("role name")

    success_count = 0
    errors = []

    roles_cache = {}
    depts_cache = {}

    default_pw_hash = get_password_hash("Password@123")

    for row_num, row in enumerate(rows[1:], start=2):
        if not any(row):  # skip completely empty row
            continue
        
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        email = str(row[email_idx]).strip() if row[email_idx] is not None else ""
        dept_name = str(row[dept_idx]).strip() if row[dept_idx] is not None else ""
        role_name = str(row[role_idx]).strip() if row[role_idx] is not None else ""

        if not name or not email or not dept_name or not role_name:
            errors.append(f"Row {row_num}: Missing required field (all fields are required)")
            continue

        if "@" not in email:
            errors.append(f"Row {row_num}: Invalid email format '{email}'")
            continue

        try:
            # 1. Resolve department (case-insensitive)
            dept_key = dept_name.lower()
            if dept_key not in depts_cache:
                dept_res = await db.execute(select(Department).where(func.lower(Department.name) == dept_key))
                dept = dept_res.scalar_one_or_none()
                if not dept:
                    code = "".join(w[0] for w in dept_name.split() if w[0].isalnum()).upper()
                    if not code:
                        code = dept_name[:3].upper()
                    
                    code_idx = 1
                    base_code = code
                    while True:
                        code_check = await db.execute(select(Department).where(Department.code == code))
                        if not code_check.scalar_one_or_none():
                            break
                        code = f"{base_code}{code_idx}"
                        code_idx += 1
                        
                    dept = Department(name=dept_name, code=code)
                    db.add(dept)
                    await db.flush()
                depts_cache[dept_key] = dept.id
            dept_id = depts_cache[dept_key]

            # 2. Resolve role (case-insensitive)
            role_key = role_name.lower()
            if role_key not in roles_cache:
                role_res = await db.execute(select(RoleManager).where(func.lower(RoleManager.name) == role_key))
                role = role_res.scalar_one_or_none()
                if not role:
                    val = role_name.lower().replace(" ", "_")
                    group_key = "verifier_general"
                    if "faculty" in val:
                        group_key = "faculty"
                    elif "head" in val or "hod" in val:
                        group_key = "hod"
                    elif "assistant" in val or "da" in val:
                        group_key = "verifier_da"
                    elif "superintendent" in val or "sp" in val:
                        group_key = "verifier_sp"
                    elif "director" in val or "registrar" in val:
                        group_key = "apex_approver"
                    elif "admin" in val:
                        group_key = "admin"
                        
                    role = RoleManager(name=role_name, value=val, group_key=group_key)
                    db.add(role)
                    await db.flush()
                roles_cache[role_key] = role.id
            role_id = roles_cache[role_key]

            # 3. Create or update user
            user_res = await db.execute(select(User).where(User.email == email.lower()))
            existing_user = user_res.scalar_one_or_none()
            if existing_user:
                existing_user.name = name
                existing_user.department_id = dept_id
                existing_user.role_id = role_id
            else:
                new_user = User(
                    name=name,
                    email=email.lower(),
                    hashed_password=default_pw_hash,
                    department_id=dept_id,
                    role_id=role_id,
                    is_active=True,
                )
                db.add(new_user)
            
            await db.flush()
            success_count += 1
        except Exception as ex:
            errors.append(f"Row {row_num}: Database error {str(ex)}")

    await db.commit()
    return {
        "success": len(errors) == 0,
        "imported": success_count,
        "errors": errors
    }


@router.put("/users/{user_id}")
async def update_user(user_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    if "name" in body:
        u.name = body["name"]
    if "email" in body:
        new_email = body["email"].lower()
        if new_email != u.email:
            existing = await db.execute(select(User).where(User.email == new_email))
            if existing.scalar_one_or_none():
                raise HTTPException(status_code=409, detail="Email already in use")
            u.email = new_email
    if "password" in body and body["password"]:
        u.hashed_password = get_password_hash(body["password"])
    if "designation" in body:
        u.designation = body["designation"]
    if "role_id" in body:
        u.role_id = body["role_id"]
    if "department_id" in body:
        u.department_id = body["department_id"]
    if "is_active" in body:
        u.is_active = bool(body["is_active"])
    if "is_approved" in body:
        u.is_approved = bool(body["is_approved"])
    await db.commit()
    return {"message": "User updated"}


@router.post("/users/{user_id}/reset-password")
async def reset_password(user_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    new_password = body.get("password", "Password@123")
    u.hashed_password = get_password_hash(new_password)
    await db.commit()
    return {"message": "Password reset successfully"}


# ─────────────────────────────────────────────────────────────────────────────
# DEPARTMENTS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/departments")
async def list_departments(db: AsyncSession = Depends(get_db), _=DeanOrAdminDep):
    result = await db.execute(select(Department).order_by(Department.short_code))
    return [{"id": d.id, "name": d.name, "short_code": d.short_code} for d in result.scalars()]


@router.post("/departments")
async def create_department(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    d = Department(name=body["name"], short_code=body["short_code"].upper())
    db.add(d)
    await db.commit()
    return {"message": "Department created", "id": d.id}


# ─────────────────────────────────────────────────────────────────────────────
# ROLES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/roles")
async def list_roles(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(RoleManager).order_by(RoleManager.name))
    return [{"id": r.id, "name": r.name, "value": r.value, "group_key": r.group_key} for r in result.scalars()]


@router.post("/roles")
async def create_role(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    name = body.get("name")
    value = body.get("value")
    group_key = body.get("group_key")
    if not name or not value or not group_key:
        raise HTTPException(status_code=400, detail="Missing name, value, or group_key")
    
    res = await db.execute(select(RoleManager).where(RoleManager.value == value))
    existing = res.scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail=f"Role value '{value}' already exists")
        
    role = RoleManager(name=name, value=value, group_key=group_key)
    db.add(role)
    await db.commit()
    return {"message": "Role created", "id": role.id}


# ─────────────────────────────────────────────────────────────────────────────
# FINANCIAL YEARS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/financial-years")
async def list_financial_years(db: AsyncSession = Depends(get_db), _=DeanOrAdminDep):
    result = await db.execute(select(FinancialYear).order_by(FinancialYear.start_date.desc()))
    return [{"id": fy.id, "label": fy.label, "start_date": fy.start_date.isoformat(), "end_date": fy.end_date.isoformat(), "is_active": fy.is_active} for fy in result.scalars()]


@router.post("/financial-years")
async def create_financial_year(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    from datetime import date as dateobj
    fy = FinancialYear(
        label=body["label"],
        start_date=dateobj.fromisoformat(body["start_date"]),
        end_date=dateobj.fromisoformat(body["end_date"]),
        is_active=body.get("is_active", True),
        is_closed=body.get("is_closed", False),
    )
    db.add(fy)
    await db.commit()
    return {"message": "Financial year created", "id": fy.id}


@router.post("/financial-years/rollover")
async def financial_year_rollover(
    db: AsyncSession = Depends(get_db),
    _=AdminDep
):
    # 1. Find the current active financial year
    active_fy_stmt = select(FinancialYear).where(FinancialYear.is_active == True)
    active_fy_res = await db.execute(active_fy_stmt)
    active_fy = active_fy_res.scalar_one_or_none()
    if not active_fy:
        raise HTTPException(status_code=400, detail="No active financial year found to rollover.")

    # 2. Parse current active FY label to calculate the next FY label
    import re
    match = re.match(r"^(\d{4})-(\d{2})$", active_fy.label)
    if not match:
        raise HTTPException(status_code=400, detail=f"Invalid financial year label format: {active_fy.label}")
    
    start_year = int(match.group(1))
    end_year_short = int(match.group(2))
    
    next_start_year = start_year + 1
    next_end_year_short = (end_year_short + 1) % 100
    next_label = f"{next_start_year}-{next_end_year_short:02d}"
    
    from datetime import date
    next_start_date = date(next_start_year, 4, 1)
    next_end_date = date(next_start_year + 1, 3, 31)

    # Find if next FY already exists
    next_fy_stmt = select(FinancialYear).where(FinancialYear.label == next_label)
    next_fy_res = await db.execute(next_fy_stmt)
    next_fy = next_fy_res.scalar_one_or_none()
    
    if next_fy:
        if next_fy.is_closed:
            raise HTTPException(status_code=400, detail=f"The next financial year '{next_label}' is already closed.")
        next_fy.is_active = True
    else:
        next_fy = FinancialYear(
            label=next_label,
            start_date=next_start_date,
            end_date=next_end_date,
            is_active=True,
            is_closed=False
        )
        db.add(next_fy)
        await db.flush()

    # Close current active FY
    active_fy.is_active = False
    active_fy.is_closed = True

    # 3. Find all active/in-progress PRs in the closed FY
    from app.models.purchase_request import PurchaseRequest, RequestStatus, PurchaseRequestItem, PurchaseRequestFlow, PurchaseRequestAssignment, TechnicalEvaluation, FinancialEvaluation, CommercialEvaluation, Document, PRReferral, PurchaseRequestHistory
    
    active_prs_res = await db.execute(
        select(PurchaseRequest)
        .options(
            selectinload(PurchaseRequest.items).selectinload(PurchaseRequestItem.budget_file),
            selectinload(PurchaseRequest.flow),
            selectinload(PurchaseRequest.assignments),
            selectinload(PurchaseRequest.technical_evaluations),
            selectinload(PurchaseRequest.financial_evaluations),
            selectinload(PurchaseRequest.commercial_evaluations),
            selectinload(PurchaseRequest.documents),
            selectinload(PurchaseRequest.referrals),
            selectinload(PurchaseRequest.initiator).selectinload(User.department)
        )
        .where(
            PurchaseRequest.financial_year_id == active_fy.id,
            PurchaseRequest.current_status.in_([
                RequestStatus.PR_SUBMITTED,
                RequestStatus.IN_PROGRESS,
                RequestStatus.SENT_BACK
            ])
        )
    )
    old_prs = active_prs_res.scalars().all()

    from app.services.budget_service import BudgetService
    budget_svc = BudgetService(db)
    dept_seqs = {}

    for old_pr in old_prs:
        # Unlock budget reservations in old FY
        await budget_svc.unlock_amount(old_pr)

        # For each item, resolve or clone its budget file to the new FY
        for item in old_pr.items:
            old_bm = item.budget_file
            if old_bm:
                new_bm_file_no = old_bm.file_no.replace(active_fy.label, next_fy.label)
                new_bm_file_no_alt = old_bm.file_no.replace(active_fy.label.lower(), next_fy.label.lower())
                
                bm_find = await db.execute(
                    select(BudgetMaster).where(
                        BudgetMaster.financial_year_id == next_fy.id,
                        BudgetMaster.file_no.in_([new_bm_file_no, new_bm_file_no_alt, old_bm.file_no])
                    )
                )
                new_bm = bm_find.scalar_one_or_none()
                if not new_bm:
                    new_bm = BudgetMaster(
                        department_id=old_bm.department_id,
                        financial_year_id=next_fy.id,
                        expenditure_category=old_bm.expenditure_category,
                        item_name=old_bm.item_name,
                        category=old_bm.category,
                        course_code=old_bm.course_code,
                        unit_cost=old_bm.unit_cost,
                        quantity=old_bm.quantity,
                        total_allocation=old_bm.total_allocation,
                        file_no=new_bm_file_no,
                        is_revision=old_bm.is_revision,
                        expert1_id=old_bm.expert1_id,
                        expert2_id=old_bm.expert2_id,
                        director_faculty_id=old_bm.director_faculty_id,
                        committed_amount=0.0,
                        utilized_amount=0.0
                    )
                    db.add(new_bm)
                    await db.flush()

        # Generate revised reference sequence number per department
        dept_id = old_pr.initiator.department_id
        if dept_id not in dept_seqs:
            q = await db.execute(
                select(func.count(PurchaseRequest.id))
                .join(User, PurchaseRequest.initiator_id == User.id)
                .where(
                    PurchaseRequest.financial_year_id == next_fy.id,
                    PurchaseRequest.parent_pr_id.isnot(None),
                    User.department_id == dept_id
                )
            )
            dept_seqs[dept_id] = q.scalar() or 0
        
        dept_seqs[dept_id] += 1
        seq_num = dept_seqs[dept_id]
        dept_code = old_pr.initiator.department.short_code if old_pr.initiator.department else "GEN"
        revised_ref = f"PR/{dept_code}/{next_fy.label}/R-{seq_num:02d}"

        # Clone the Purchase Request
        new_pr = PurchaseRequest(
            category_id=old_pr.category_id,
            financial_year_id=next_fy.id,
            initiator_id=old_pr.initiator_id,
            nominee_id=old_pr.nominee_id,
            procurement_id=old_pr.procurement_id,
            purchase_type=old_pr.purchase_type,
            amount=old_pr.amount,
            emd=old_pr.emd,
            performance_security=old_pr.performance_security,
            current_status=old_pr.current_status,
            vendor_list_link=old_pr.vendor_list_link,
            is_item_split=old_pr.is_item_split,
            item_split_justification=old_pr.item_split_justification,
            is_quantity_split=old_pr.is_quantity_split,
            quantity_split_details=old_pr.quantity_split_details,
            is_service_center_in_south=old_pr.is_service_center_in_south,
            service_center_south_desc=old_pr.service_center_south_desc,
            basis_of_estimate_details=old_pr.basis_of_estimate_details,
            delivery_mode=old_pr.delivery_mode,
            delivery_location=old_pr.delivery_location,
            exemption=old_pr.exemption,
            exemption_remarks=old_pr.exemption_remarks,
            is_training_required=old_pr.is_training_required,
            training_type=old_pr.training_type,
            training_vendor=old_pr.training_vendor,
            training_comments=old_pr.training_comments,
            tender_reference_number=old_pr.tender_reference_number,
            date_of_tender=old_pr.date_of_tender,
            date_of_tech_bid_opening=old_pr.date_of_tech_bid_opening,
            date_of_financial_bid_opening=old_pr.date_of_financial_bid_opening,
            aa_approved_at=old_pr.aa_approved_at,
            te_initiated_at=old_pr.te_initiated_at,
            te_approved_at=old_pr.te_approved_at,
            fs_initiated_at=old_pr.fs_initiated_at,
            fs_approved_at=old_pr.fs_approved_at,
            po_initiated_at=old_pr.po_initiated_at,
            po_approved_at=old_pr.po_approved_at,
            faculty1_id=old_pr.faculty1_id,
            faculty2_id=old_pr.faculty2_id,
            faculty3_id=old_pr.faculty3_id,
            aa_approver_id=old_pr.aa_approver_id,
            form_data=old_pr.form_data,
            parent_pr_id=old_pr.id,
            lpc_remarks=old_pr.lpc_remarks,
            lpc_committee_members=old_pr.lpc_committee_members,
            lpc_minutes_reference=old_pr.lpc_minutes_reference,
            single_bid_justification=old_pr.single_bid_justification,
            icr_number=revised_ref
        )
        db.add(new_pr)
        await db.flush()

        # Clone items, linking to new budgets
        for item in old_pr.items:
            old_item_bm = item.budget_file
            new_item_bm = None
            if old_item_bm:
                new_item_bm_file_no = old_item_bm.file_no.replace(active_fy.label, next_fy.label)
                new_item_bm_file_no_alt = old_item_bm.file_no.replace(active_fy.label.lower(), next_fy.label.lower())
                new_bm_find = await db.execute(
                    select(BudgetMaster).where(
                        BudgetMaster.financial_year_id == next_fy.id,
                        BudgetMaster.file_no.in_([new_item_bm_file_no, new_item_bm_file_no_alt, old_item_bm.file_no])
                    )
                )
                new_item_bm = new_bm_find.scalar_one_or_none()

            new_item = PurchaseRequestItem(
                purchase_request_id=new_pr.id,
                budget_file_id=new_item_bm.id if new_item_bm else None,
                item_description=item.item_description,
                quantity=item.quantity,
                estimated_total=item.estimated_total,
                charges=item.charges,
                requirement_type=item.requirement_type,
                availability=item.availability,
                availability_remarks=item.availability_remarks,
                site_readiness=item.site_readiness,
                site_readiness_remarks=item.site_readiness_remarks,
                warranty=item.warranty,
                delivery_period=item.delivery_period,
                present_stock=item.present_stock,
                justification_for_procurement=item.justification_for_procurement,
                previous_file_no_reference=item.previous_file_no_reference,
                installation_required=item.installation_required,
                tech_specs_text=item.tech_specs_text,
                gem_link=item.gem_link,
            )
            db.add(new_item)
        await db.flush()

        # Clone flow state
        if old_pr.flow:
            new_flow = PurchaseRequestFlow(
                purchase_request_id=new_pr.id,
                phase_id=old_pr.flow.phase_id,
                step_order=old_pr.flow.step_order,
                rejected=old_pr.flow.rejected
            )
            db.add(new_flow)

        # Clone assignments
        for ass in old_pr.assignments:
            new_ass = PurchaseRequestAssignment(
                purchase_request_id=new_pr.id,
                assigned_by_id=ass.assigned_by_id,
                assigned_da_id=ass.assigned_da_id,
                status=ass.status,
                assigned_at=ass.assigned_at
            )
            db.add(new_ass)

        # Clone evaluations
        for te in old_pr.technical_evaluations:
            new_te = TechnicalEvaluation(
                purchase_request_id=new_pr.id,
                vendor_name=te.vendor_name,
                is_qualified=te.is_qualified,
                remarks=te.remarks,
                created_at=te.created_at
            )
            db.add(new_te)
        
        for fe in old_pr.financial_evaluations:
            new_fe = FinancialEvaluation(
                purchase_request_id=new_pr.id,
                vendor_name=fe.vendor_name,
                quoted_amount=fe.quoted_amount,
                ranking=fe.ranking,
                is_awarded=fe.is_awarded,
                remarks=fe.remarks,
                unit_price=fe.unit_price,
                taxes=fe.taxes,
                delivery_period=fe.delivery_period,
                warranty=fe.warranty,
                created_at=fe.created_at
            )
            db.add(new_fe)
            
        for ce in old_pr.commercial_evaluations:
            new_ce = CommercialEvaluation(
                purchase_request_id=new_pr.id,
                vendor_name=ce.vendor_name,
                vendor_email=ce.vendor_email,
                quoted_amount=ce.quoted_amount,
                is_qualified=ce.is_qualified,
                remarks=ce.remarks
            )
            db.add(new_ce)

        # Clone documents
        for doc in old_pr.documents:
            new_doc = Document(
                purchase_request_id=new_pr.id,
                doc_key=doc.doc_key,
                doc_value=doc.doc_value,
                uploaded_by_id=doc.uploaded_by_id,
                updated_at=doc.updated_at
            )
            db.add(new_doc)

        # Clone referrals
        for ref in old_pr.referrals:
            new_ref = PRReferral(
                purchase_request_id=new_pr.id,
                referred_by_id=ref.referred_by_id,
                referred_to_id=ref.referred_to_id,
                query=ref.query,
                query_document_path=ref.query_document_path,
                response=ref.response,
                response_document_path=ref.response_document_path,
                status=ref.status,
                created_at=ref.created_at,
                responded_at=ref.responded_at
            )
            db.add(new_ref)

        # Lock budget amounts on the new cloned budget files
        await budget_svc.lock_amount(new_pr)

        # Mark old PR as rolled_over
        old_pr.current_status = RequestStatus.ROLLED_OVER

        # Add history entries
        old_history = PurchaseRequestHistory(
            purchase_request_id=old_pr.id,
            status="Rolled Over",
            remarks=f"Purchase request rolled over to next financial year {next_fy.label}. Revised reference: {revised_ref}",
            acted_at=datetime.utcnow()
        )
        db.add(old_history)

        new_history = PurchaseRequestHistory(
            purchase_request_id=new_pr.id,
            status="Rolled Over",
            remarks=f"Purchase request rolled over from financial year {active_fy.label}. Original reference: {old_pr.icr_number}",
            acted_at=datetime.utcnow()
        )
        db.add(new_history)

    await db.commit()
    return {
        "message": "Financial Year rollover completed successfully.",
        "closed_year": active_fy.label,
        "opened_year": next_fy.label,
        "rolled_over_count": len(old_prs)
    }


# ─────────────────────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/settings")
async def get_settings(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(Settings))
    return {s.key_name: s.value for s in result.scalars()}


@router.put("/settings/{key}")
async def update_setting(key: str, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(Settings).where(Settings.key_name == key))
    setting = result.scalar_one_or_none()
    if setting:
        setting.value = body["value"]
    else:
        setting = Settings(key_name=key, value=body["value"])
        db.add(setting)
    await db.commit()
    return {"message": "Setting updated"}


# ─────────────────────────────────────────────────────────────────────────────
# BUDGET
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/budget")
async def list_budget(
    skip: int = 0,
    limit: int = Query(default=50, le=200),
    db: AsyncSession = Depends(get_db),
    _=BudgetViewDep
):
    # Get total count
    total = await db.scalar(select(func.count(BudgetMaster.id))) or 0

    result = await db.execute(
        select(BudgetMaster)
        .outerjoin(Department)
        .options(
            selectinload(BudgetMaster.expert1),
            selectinload(BudgetMaster.expert2),
            selectinload(BudgetMaster.director_faculty)
        )
        .order_by(Department.short_code.asc(), BudgetMaster.file_no.asc())
        .offset(skip)
        .limit(limit)
    )
    entries = result.scalars().all()
    items = [
        {
            "id": b.id, "item_name": b.item_name,
            "total_cost": b.total_allocation,
            "total_allocation": b.total_allocation,
            "locked_amount": b.committed_amount,
            "committed_amount": b.committed_amount,
            "deducted_amount": b.utilized_amount,
            "utilized_amount": b.utilized_amount,
            "available_amount": b.available_balance,
            "available_balance": b.available_balance,
            "department_id": b.department_id,
            "financial_year_id": b.financial_year_id, "expenditure_category": b.expenditure_category,
            "category": b.category, "unit_cost": b.unit_cost, "quantity": b.quantity,
            "file_no": b.file_no,
            "expert1_id": b.expert1_id,
            "expert2_id": b.expert2_id,
            "director_faculty_id": b.director_faculty_id,
            "expert1": {"id": b.expert1.id, "name": b.expert1.name, "email": b.expert1.email} if b.expert1 else None,
            "expert2": {"id": b.expert2.id, "name": b.expert2.name, "email": b.expert2.email} if b.expert2 else None,
            "director_faculty": {"id": b.director_faculty.id, "name": b.director_faculty.name, "email": b.director_faculty.email} if b.director_faculty else None,
        }
        for b in entries
    ]
    return {"items": items, "total": total}


@router.get("/budget/summary")
async def budget_summary(db: AsyncSession = Depends(get_db), _=DeanOrAdminDep):
    """System-wide consolidated budget totals for admin dashboard."""
    result = await db.execute(select(BudgetMaster))
    entries = result.scalars().all()
    total = sum(b.total_cost for b in entries)
    locked = sum(b.locked_amount for b in entries)
    deducted = sum(b.deducted_amount for b in entries)
    return {"total": total, "locked": locked, "deducted": deducted, "available": total - locked - deducted}


async def get_stored_categories(db: AsyncSession):
    result_exp = await db.execute(select(Settings).where(Settings.key_name == "budget_expenditure_categories"))
    exp_setting = result_exp.scalar_one_or_none()
    if exp_setting:
        exp_list = [c.strip() for c in exp_setting.value.split(",") if c.strip()]
    else:
        exp_list = ["CAPEX", "OPEX"]
        db.add(Settings(key_name="budget_expenditure_categories", value="CAPEX,OPEX"))
        await db.flush()

    result_item = await db.execute(select(Settings).where(Settings.key_name == "budget_item_categories"))
    item_setting = result_item.scalar_one_or_none()
    if item_setting:
        item_list = [c.strip() for c in item_setting.value.split(",") if c.strip()]
    else:
        item_list = ["computer", "lab_equipment", "software", "furniture"]
        db.add(Settings(key_name="budget_item_categories", value="computer,lab_equipment,software,furniture"))
        await db.flush()

    # Also load added_by_dean
    result_dean = await db.execute(select(Settings).where(Settings.key_name == "budget_categories_added_by_dean"))
    dean_setting = result_dean.scalar_one_or_none()
    if dean_setting:
        import json
        try:
            dean_cats = json.loads(dean_setting.value)
        except Exception:
            dean_cats = {"expenditure": [], "item": []}
    else:
        dean_cats = {"expenditure": [], "item": []}

    return {
        "expenditure_categories": exp_list,
        "item_categories": item_list,
        "added_by_dean": dean_cats
    }


@router.get("/budget/categories")
async def get_budget_categories(db: AsyncSession = Depends(get_db), _=BudgetViewDep):
    return await get_stored_categories(db)


@router.post("/budget/categories")
async def add_budget_category(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    await db.refresh(current_user, ["role"])
    if current_user.role.group_key not in ["admin", "dean_approver", "apex_approver"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    cat_type = body.get("type")
    val = body.get("value")
    if not val or not val.strip():
        raise HTTPException(status_code=400, detail="Category value is required")
    val = val.strip()

    if cat_type == "expenditure":
        key = "budget_expenditure_categories"
    elif cat_type == "item":
        key = "budget_item_categories"
    else:
        raise HTTPException(status_code=400, detail="Invalid type: must be 'expenditure' or 'item'")

    result = await db.execute(select(Settings).where(Settings.key_name == key))
    setting = result.scalar_one_or_none()
    
    already_exists = False
    
    if setting:
        existing = [c.strip() for c in setting.value.split(",") if c.strip()]
        if val in existing:
            already_exists = True
        else:
            existing.append(val)
            setting.value = ",".join(existing)
    else:
        if cat_type == "expenditure":
            defaults = ["CAPEX", "OPEX", val]
        else:
            defaults = ["computer", "lab_equipment", "software", "furniture", val]
        setting = Settings(key_name=key, value=",".join(defaults))
        db.add(setting)

    # Record if added by dean budget role
    if not already_exists and current_user.role.group_key == "dean_approver":
        result_dean = await db.execute(select(Settings).where(Settings.key_name == "budget_categories_added_by_dean"))
        dean_setting = result_dean.scalar_one_or_none()
        import json
        if dean_setting:
            try:
                dean_cats = json.loads(dean_setting.value)
            except Exception:
                dean_cats = {"expenditure": [], "item": []}
            
            if cat_type == "expenditure":
                if "expenditure" not in dean_cats:
                    dean_cats["expenditure"] = []
                if val not in dean_cats["expenditure"]:
                    dean_cats["expenditure"].append(val)
            else:
                if "item" not in dean_cats:
                    dean_cats["item"] = []
                if val not in dean_cats["item"]:
                    dean_cats["item"].append(val)
            dean_setting.value = json.dumps(dean_cats)
        else:
            if cat_type == "expenditure":
                dean_cats = {"expenditure": [val], "item": []}
            else:
                dean_cats = {"expenditure": [], "item": [val]}
            db.add(Settings(key_name="budget_categories_added_by_dean", value=json.dumps(dean_cats)))

    await db.commit()
    return await get_stored_categories(db)


@router.delete("/budget/categories")
async def delete_budget_category(
    type: str = Query(...),
    value: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    await db.refresh(current_user, ["role"])
    if current_user.role.group_key != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete categories")

    value = value.strip()
    if not value:
        raise HTTPException(status_code=400, detail="Category value is required")

    if type == "expenditure":
        key = "budget_expenditure_categories"
    elif type == "item":
        key = "budget_item_categories"
    else:
        raise HTTPException(status_code=400, detail="Invalid type: must be 'expenditure' or 'item'")

    # Verify if category was added by dean budget role
    result_dean = await db.execute(select(Settings).where(Settings.key_name == "budget_categories_added_by_dean"))
    dean_setting = result_dean.scalar_one_or_none()
    is_dean_added = False
    dean_cats = {"expenditure": [], "item": []}
    
    if dean_setting:
        import json
        try:
            dean_cats = json.loads(dean_setting.value)
        except Exception:
            dean_cats = {"expenditure": [], "item": []}
        
        if type == "expenditure" and value in dean_cats.get("expenditure", []):
            is_dean_added = True
        elif type == "item" and value in dean_cats.get("item", []):
            is_dean_added = True

    if not is_dean_added:
        raise HTTPException(status_code=400, detail="Category cannot be deleted: it was not added by the dean budget role")

    # Check if category is currently used by any BudgetMaster entry
    if type == "expenditure":
        count_stmt = select(func.count(BudgetMaster.id)).where(BudgetMaster.expenditure_category == value)
    else:
        count_stmt = select(func.count(BudgetMaster.id)).where(BudgetMaster.category == value)

    count_res = await db.execute(count_stmt)
    use_count = count_res.scalar() or 0
    if use_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Category '{value}' is in use by {use_count} budget file(s) and cannot be deleted"
        )

    # Perform deletion from Settings
    result_setting = await db.execute(select(Settings).where(Settings.key_name == key))
    setting = result_setting.scalar_one_or_none()
    if setting:
        existing = [c.strip() for c in setting.value.split(",") if c.strip()]
        if value in existing:
            existing.remove(value)
            setting.value = ",".join(existing)

    # Remove from dean_cats metadata list
    if type == "expenditure":
        if value in dean_cats.get("expenditure", []):
            dean_cats["expenditure"].remove(value)
    else:
        if value in dean_cats.get("item", []):
            dean_cats["item"].remove(value)

    if dean_setting:
        dean_setting.value = json.dumps(dean_cats)

    await db.commit()
    return await get_stored_categories(db)


@router.get("/budget/next-file-number")
async def get_next_file_number(
    department_id: int,
    expenditure_category: str,
    financial_year_id: int,
    db: AsyncSession = Depends(get_db),
    _=BudgetViewDep
):
    from app.models.user import Department
    from app.models.budget import FinancialYear, BudgetMaster
    from sqlalchemy import func

    dept_res = await db.execute(select(Department).where(Department.id == department_id))
    dept = dept_res.scalar_one_or_none()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == financial_year_id))
    fy = fy_res.scalar_one_or_none()
    if not fy:
        raise HTTPException(status_code=404, detail="Financial Year not found")

    stmt = select(func.count(BudgetMaster.id)).where(
        and_(
            BudgetMaster.department_id == department_id,
            BudgetMaster.expenditure_category == expenditure_category,
            BudgetMaster.financial_year_id == financial_year_id
        )
    )
    count_res = await db.execute(stmt)
    count = count_res.scalar() or 0
    next_num = count + 1

    dept_code = dept.short_code.lower()
    source_code = expenditure_category.lower()
    fy_label = fy.label.lower()

    file_no = f"nitt/{dept_code}/{source_code}/{fy_label}/{next_num}"
    return {"file_no": file_no}


@router.get("/budget/{b_id}")
async def get_budget_detail(b_id: int, db: AsyncSession = Depends(get_db), _=BudgetViewDep):
    result = await db.execute(
        select(BudgetMaster).where(BudgetMaster.id == b_id)
    )
    b = result.scalar_one_or_none()
    if not b:
        raise HTTPException(status_code=404, detail="Budget not found")
    return {
        "id": b.id,
        "department_id": b.department_id,
        "financial_year_id": b.financial_year_id,
        "expenditure_category": b.expenditure_category,
        "item_name": b.item_name,
        "category": b.category,
        "unit_cost": b.unit_cost,
        "quantity": b.quantity,
        "total_cost": b.total_allocation,
        "total_allocation": b.total_allocation,
        "file_no": b.file_no,
        "expert1_id": b.expert1_id,
        "expert2_id": b.expert2_id,
        "director_faculty_id": b.director_faculty_id
    }


@router.post("/budget")
async def create_budget(body: dict, db: AsyncSession = Depends(get_db), _=DeanBudgetDep):
    fy_id = int(body["financial_year_id"])
    fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == fy_id))
    fy = fy_res.scalar_one_or_none()
    if not fy:
        raise HTTPException(status_code=400, detail="Financial Year not found")
    if fy.is_closed:
        raise HTTPException(status_code=400, detail="The selected financial year is closed. Budgets in closed financial years cannot be modified.")

    b = BudgetMaster(
        department_id=int(body["department_id"]),
        financial_year_id=fy_id,
        expenditure_category=body["expenditure_category"],
        item_name=body["item_name"],
        category=body["category"],
        course_code=body.get("course_code", "N/A"),
        unit_cost=float(body["unit_cost"]),
        quantity=int(body["quantity"]),
        total_cost=float(body["unit_cost"]) * int(body["quantity"]),
        file_no=body["file_no"],
        is_revision=False,
    )
    db.add(b)
    await db.commit()
    return {"message": "Budget created", "id": b.id}


@router.put("/budget/{b_id}")
async def update_budget(b_id: int, body: dict, db: AsyncSession = Depends(get_db), _=DeanBudgetDep):

    result = await db.execute(select(BudgetMaster).where(BudgetMaster.id == b_id))
    b = result.scalar_one_or_none()
    if not b:
        raise HTTPException(status_code=404, detail="Not found")

    # Check if current budget file belongs to closed financial year
    current_fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == b.financial_year_id))
    current_fy = current_fy_res.scalar_one_or_none()
    if current_fy and current_fy.is_closed:
        raise HTTPException(status_code=400, detail="The current financial year for this budget is closed. Budgets in closed financial years cannot be modified.")

    b.item_name = body.get("item_name", b.item_name)
    if "department_id" in body:
        b.department_id = int(body["department_id"])
    if "financial_year_id" in body:
        new_fy_id = int(body["financial_year_id"])
        new_fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == new_fy_id))
        new_fy = new_fy_res.scalar_one_or_none()
        if not new_fy:
            raise HTTPException(status_code=400, detail="Financial Year not found")
        if new_fy.is_closed:
            raise HTTPException(status_code=400, detail="The target financial year is closed. Budgets in closed financial years cannot be modified.")
        b.financial_year_id = new_fy_id
    if "expenditure_category" in body:
        b.expenditure_category = body["expenditure_category"]
    if "category" in body:
        b.category = body["category"]
    if "file_no" in body:
        b.file_no = body["file_no"]
    if "unit_cost" in body and "quantity" in body:
        b.unit_cost = float(body["unit_cost"])
        b.quantity = int(body["quantity"])
        b.total_cost = b.unit_cost * b.quantity
    await db.commit()
    return {"message": "Budget updated"}


@router.delete("/budget/clear")
async def clear_all_budgets(db: AsyncSession = Depends(get_db), _=DeanBudgetDep):
    """Deletes all budget master entries to start fresh. Skips any budgets linked to active PR items or belonging to closed financial years."""
    from app.models.purchase_request import PurchaseRequestItem
    from sqlalchemy import delete

    # Get linked budget IDs
    linked_res = await db.execute(select(PurchaseRequestItem.budget_file_id))
    linked_ids = {row[0] for row in linked_res.fetchall() if row[0] is not None}

    closed_fy_res = await db.execute(select(FinancialYear.id).where(FinancialYear.is_closed == True))
    closed_fy_ids = [row[0] for row in closed_fy_res.fetchall()]

    stmt = delete(BudgetMaster)
    if closed_fy_ids:
        stmt = stmt.where(BudgetMaster.financial_year_id.not_in(closed_fy_ids))
    if linked_ids:
        stmt = stmt.where(BudgetMaster.id.not_in(linked_ids))

    result = await db.execute(stmt)
    await db.commit()
    return {"message": f"Cleared {result.rowcount} unlinked budget files. Budgets in closed financial years and those linked to active Purchase Requests could not be deleted."}


# ─────────────────────────────────────────────────────────────────────────────
# PROCUREMENT METHODS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/procurement-methods")
async def list_procurement_methods(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(ProcurementManager).order_by(ProcurementManager.name))
    return [
        {
            "id": p.id,
            "name": p.name,
            "description": p.description,
            "max_amount": p.max_amount,
            "form_schema": p.form_schema
        }
        for p in result.scalars()
    ]


@router.post("/procurement-methods")
async def create_procurement_method(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    p = ProcurementManager(
        name=body["name"],
        description=body.get("description"),
        max_amount=float(body["max_amount"]) if body.get("max_amount") is not None else None,
        form_schema=body.get("form_schema")
    )
    db.add(p)
    await db.commit()
    return {"message": "Procurement method created", "id": p.id}


@router.put("/procurement-methods/{pm_id}")
async def update_procurement_method(pm_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(ProcurementManager).where(ProcurementManager.id == pm_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Procurement method not found")
    if "name" in body:
        p.name = body["name"]
    if "description" in body:
        p.description = body["description"]
    if "max_amount" in body:
        p.max_amount = float(body["max_amount"]) if body["max_amount"] is not None else None
    if "form_schema" in body:
        p.form_schema = body["form_schema"]
    await db.commit()
    return {"message": "Procurement method updated"}


@router.delete("/procurement-methods/{pm_id}")
async def delete_procurement_method(pm_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(ProcurementManager).where(ProcurementManager.id == pm_id))
    p = result.scalar_one_or_none()
    if p:
        await db.delete(p)
        await db.commit()
    return {"message": "Procurement method deleted"}


# ─────────────────────────────────────────────────────────────────────────────
# WORKFLOWS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/workflows")
async def list_workflows(db: AsyncSession = Depends(get_db), _=AdminDep):
    from sqlalchemy.orm import selectinload
    result = await db.execute(
        select(WorkFlowHierarchy).options(
            selectinload(WorkFlowHierarchy.role),
            selectinload(WorkFlowHierarchy.user)
        ).order_by(
            WorkFlowHierarchy.category_id,
            WorkFlowHierarchy.procurement_id,
            WorkFlowHierarchy.phase_id,
            WorkFlowHierarchy.step_order,
        )
    )
    entries = result.scalars().all()
    return [
        {
            "id": w.id,
            "category_id": w.category_id,
            "procurement_id": w.procurement_id,
            "phase_id": w.phase_id,
            "step_order": w.step_order,
            "user_type": w.user_type,
            "user_id": w.user_id,
            "user_name": w.user.name if w.user else None,
            "user_group": w.user_group,
            "role_id": w.role_id,
            "role_name": w.role.name if w.role else (w.user_group.replace("_", " ").title() if w.user_group else None),
            "is_enabled": w.is_enabled,
            "purchase_type": w.purchase_type,
            "tender_vendors_threshold": w.tender_vendors_threshold,
            "tender_vendors_comparison": w.tender_vendors_comparison,
            "skip_condition": w.skip_condition,
            "condition_field": w.condition_field,
            "condition_operator": w.condition_operator,
            "condition_value": w.condition_value,
        }
        for w in entries
    ]


@router.post("/workflows")
async def create_workflow(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    role_id = body.get("role_id")
    user_id = body.get("user_id")
    user_group = body.get("user_group")
    user_type = body.get("user_type", "verifier")

    if user_type == "user" and user_id:
        user_res = await db.execute(select(User).where(User.id == user_id))
        user_obj = user_res.scalar_one_or_none()
        if not user_obj:
            raise HTTPException(status_code=400, detail="User not found")
        role_id = None
        user_group = None
    elif user_type in ["purchase_initiator", "da_assigner", "verifier_da", "tech_evaluation"]:
        user_id = None
        role_id = None
        user_group = None
    else:
        if user_type not in ["verifier", "approver", "partial_approver"]:
            user_type = "verifier"
        user_id = None
        if role_id:
            role_res = await db.execute(select(RoleManager).where(RoleManager.id == role_id))
            role_obj = role_res.scalar_one_or_none()
            if role_obj:
                user_group = role_obj.group_key
                
    purchase_type = body.get("purchase_type", "department")
    wf = WorkFlowHierarchy(
        category_id=body["category_id"],
        phase_id=body["phase_id"],
        procurement_id=body["procurement_id"],
        step_order=body["step_order"],
        user_type=user_type,
        user_id=user_id,
        user_group=user_group,
        role_id=role_id,
        purchase_type=purchase_type,
        is_enabled=True,
        tender_vendors_threshold=body.get("tender_vendors_threshold"),
        tender_vendors_comparison=body.get("tender_vendors_comparison"),
        skip_condition=body.get("skip_condition"),
        condition_field=body.get("condition_field"),
        condition_operator=body.get("condition_operator"),
        condition_value=body.get("condition_value"),
    )
    db.add(wf)
    await db.commit()
    return {"message": "Workflow created", "id": wf.id}


@router.put("/workflows/{wf_id}")
async def update_workflow(wf_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == wf_id))
    wf = result.scalar_one_or_none()
    if not wf:
        raise HTTPException(status_code=404, detail="Not found")
    if "step_order" in body:
        wf.step_order = body["step_order"]
    if "user_type" in body:
        wf.user_type = body["user_type"]
        if wf.user_type in ["purchase_initiator", "da_assigner", "verifier_da", "tech_evaluation"]:
            wf.user_id = None
            wf.role_id = None
            wf.user_group = None
        elif wf.user_type == "user":
            wf.role_id = None
            wf.user_group = None
        elif wf.user_type not in ["verifier", "approver", "partial_approver"]:
            pass
    if "user_id" in body:
        wf.user_id = body["user_id"]
        if wf.user_id:
            wf.role_id = None
            wf.user_group = None
            wf.user_type = "user"
    if "role_id" in body:
        wf.role_id = body["role_id"]
        if wf.role_id:
            role_res = await db.execute(select(RoleManager).where(RoleManager.id == wf.role_id))
            role_obj = role_res.scalar_one_or_none()
            if role_obj:
                wf.user_group = role_obj.group_key
                wf.user_id = None
                if wf.user_type not in ["verifier", "approver", "partial_approver"]:
                    wf.user_type = "verifier"
        else:
            wf.role_id = None
    elif "user_group" in body:
        wf.user_group = body["user_group"]
        wf.user_id = None
        wf.role_id = None
        if wf.user_type not in ["verifier", "approver", "partial_approver"]:
            wf.user_type = "verifier"
    if "is_enabled" in body:
        wf.is_enabled = bool(body["is_enabled"])
    if "tender_vendors_threshold" in body:
        # Accept null/None to clear the threshold
        wf.tender_vendors_threshold = body["tender_vendors_threshold"]
    if "tender_vendors_comparison" in body:
        wf.tender_vendors_comparison = body["tender_vendors_comparison"]
    if "skip_condition" in body:
        wf.skip_condition = body["skip_condition"]
    if "condition_field" in body:
        wf.condition_field = body["condition_field"]
    if "condition_operator" in body:
        wf.condition_operator = body["condition_operator"]
    if "condition_value" in body:
        wf.condition_value = body["condition_value"]
    await db.commit()
    return {"message": "Workflow updated"}


@router.post("/workflows/reorder")
async def reorder_workflows(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    steps = body.get("steps", [])
    for item in steps:
        res = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == item["id"]))
        wf = res.scalar_one_or_none()
        if wf:
            wf.step_order = item["step_order"]
    await db.commit()
    return {"message": "Steps reordered"}


@router.patch("/workflows/{wf_id}/toggle")
async def toggle_workflow_step(wf_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == wf_id))
    wf = result.scalar_one_or_none()
    if not wf:
        raise HTTPException(status_code=404, detail="Not found")
    wf.is_enabled = not wf.is_enabled
    await db.commit()
    return {"message": "Toggled", "is_enabled": wf.is_enabled}


@router.delete("/workflows/{wf_id}")
async def delete_workflow(wf_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(WorkFlowHierarchy).where(WorkFlowHierarchy.id == wf_id))
    wf = result.scalar_one_or_none()
    if wf:
        await db.delete(wf)
        await db.commit()
    return {"message": "Workflow deleted"}


@router.post("/workflows/reset-defaults")
async def reset_workflows(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    """Reset workflow steps for one category × procurement × purchase type to seeded defaults."""
    from sqlalchemy import delete
    from app.seed_workflows import build_workflow_steps

    cat_id = body.get("category_id")
    proc_id = body.get("procurement_id")
    purchase_type = body.get("purchase_type", "department")
    if not cat_id or not proc_id:
        raise HTTPException(status_code=400, detail="Missing category_id or procurement_id")

    cat_res = await db.execute(select(PurchaseCategory).where(PurchaseCategory.id == cat_id))
    cat = cat_res.scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")

    proc_res = await db.execute(select(ProcurementManager).where(ProcurementManager.id == proc_id))
    proc = proc_res.scalar_one_or_none()
    if not proc:
        raise HTTPException(status_code=404, detail="Procurement method not found")

    await db.execute(
        delete(WorkFlowHierarchy).where(
            WorkFlowHierarchy.category_id == cat_id,
            WorkFlowHierarchy.procurement_id == proc_id,
            WorkFlowHierarchy.purchase_type == purchase_type,
        )
    )

    roles_res = await db.execute(select(RoleManager))
    roles = {r.value: r for r in roles_res.scalars()}

    phases_res = await db.execute(select(PhaseManager))
    phases = {}
    for p in phases_res.scalars():
        key = {"Administrative Approval": "AA", "Tendering": "TD", "Technical Evaluation": "TE",
               "Financial Sanction": "FS", "Purchase Order": "PO"}.get(p.phase_name)
        if key:
            phases[key] = p

    cat_key = "cat1" if cat.max_amount <= 100_000 else ("cat2" if cat.max_amount <= 1_000_000 else "cat3")
    categories = {cat_key: cat}

    all_rows = build_workflow_steps(roles, phases, categories, [proc])
    for w in all_rows:
        if w.purchase_type == purchase_type:
            db.add(w)

    await db.commit()
    return {"message": "Workflows reset to defaults"}



# ─────────────────────────────────────────────────────────────────────────────
# PHASES & CATEGORIES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/phases")
async def list_phases(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(PhaseManager).order_by(PhaseManager.phase_order))
    return [{"id": p.id, "phase_name": p.phase_name} for p in result.scalars()]


@router.get("/categories")
async def list_categories(procurement_id: Optional[int] = None, db: AsyncSession = Depends(get_db), _=AdminDep):
    stmt = select(PurchaseCategory)
    if procurement_id is not None:
        stmt = stmt.where(PurchaseCategory.procurement_id == procurement_id)
    stmt = stmt.order_by(PurchaseCategory.procurement_id, PurchaseCategory.min_amount)
    result = await db.execute(stmt)
    return [
        {
            "id": c.id,
            "title": c.title,
            "min_amount": c.min_amount,
            "max_amount": c.max_amount,
            "is_active": c.is_active,
            "procurement_id": c.procurement_id,
            "requirement_type": c.requirement_type,
        }
        for c in result.scalars()
    ]


@router.post("/categories")
async def create_category(body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    c = PurchaseCategory(
        title=body["title"],
        min_amount=float(body["min_amount"]),
        max_amount=float(body["max_amount"]),
        is_active=body.get("is_active", True),
        procurement_id=int(body["procurement_id"]),
        requirement_type=body.get("requirement_type") if body.get("requirement_type") else None
    )
    db.add(c)
    await db.commit()
    return {"message": "Category created", "id": c.id}


@router.put("/categories/{cat_id}")
async def update_category(cat_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(PurchaseCategory).where(PurchaseCategory.id == cat_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    if "title" in body:
        c.title = body["title"]
    if "min_amount" in body:
        c.min_amount = float(body["min_amount"])
    if "max_amount" in body:
        c.max_amount = float(body["max_amount"])
    if "is_active" in body:
        c.is_active = bool(body["is_active"])
    if "procurement_id" in body:
        c.procurement_id = int(body["procurement_id"])
    if "requirement_type" in body:
        c.requirement_type = body["requirement_type"] if body["requirement_type"] else None
    await db.commit()
    return {"message": "Category updated"}


@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(PurchaseCategory).where(PurchaseCategory.id == cat_id))
    c = result.scalar_one_or_none()
    if c:
        await db.delete(c)
        await db.commit()
    return {"message": "Category deleted"}


# ─────────────────────────────────────────────────────────────────────────────
# PENDING USER ONBOARDING
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/users/pending")
async def get_pending_users(db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(
        select(User)
        .options(selectinload(User.role), selectinload(User.department))
        .where(User.is_approved == False)
        .order_by(User.created_at.desc())
    )
    users = result.scalars().all()
    return [
        {
            "id": u.id,
            "name": u.name,
            "email": u.email,
            "designation": u.designation,
            "gender": u.gender,
            "role": {"group_key": u.role.group_key, "name": u.role.name} if u.role else None,
            "department": {"name": u.department.name, "short_code": u.department.short_code} if u.department else None,
            "signature_path": f"/storage/{u.signature_path}" if u.signature_path else None,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        }
        for u in users
    ]


@router.post("/users/{user_id}/approve")
async def approve_user(user_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    u.is_approved = True
    await db.commit()
    return {"message": "User approved successfully"}


@router.post("/users/{user_id}/reject")
async def reject_user(user_id: int, db: AsyncSession = Depends(get_db), _=AdminDep):
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    if u.signature_path:
        import os
        abs_path = os.path.join(settings.STORAGE_PATH, u.signature_path)
        if os.path.exists(abs_path):
            try:
                os.remove(abs_path)
            except Exception:
                pass
    await db.delete(u)
    await db.commit()
    return {"message": "User onboarding request rejected and deleted"}


# ─────────────────────────────────────────────────────────────────────────────
# BUDGET CSV EXPORT & IMPORT
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/budget/import-template")
async def download_budget_template(_=DeanBudgetDep):
    import io
    import csv
    from fastapi.responses import StreamingResponse

    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["S.No", "Department", "File No", "Procurement", "Budget Amount (INR)"])
    writer.writerow(["1", "CSE", "NITT/CSE/2026-27/005", "GPU Server Purchase", "1500000"])
    writer.writerow(["2", "ECE", "NITT/ECE/2026-27/001", "Lab Oscilloscopes", "800000"])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="budget_import_template.csv"'}
    )


@router.post("/budget/import")
@limiter.limit("10/minute")
async def import_budget_csv(
    request: Request,
    file: UploadFile,
    financial_year_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    _=DeanBudgetDep
):
    import io
    import csv
    import re
    from app.models.user import Department
    from app.models.budget import BudgetMaster, FinancialYear
    from sqlalchemy import select, and_

    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files (.csv) are supported")

    content = await file.read()
    try:
        csv_text = content.decode("utf-8")
        reader = csv.reader(io.StringIO(csv_text))
        rows = list(reader)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="The CSV file is empty")

    # Normalize headers
    normalized_headers = [h.strip().lower().replace("_", " ") for h in rows[0]]

    # Helper to find index of a column matching a list of possible substring/patterns
    def find_idx(keywords, required=False):
        for i, h in enumerate(normalized_headers):
            if any(kw in h for kw in keywords):
                return i
        if required:
            raise HTTPException(status_code=400, detail=f"Required column matching one of {keywords} not found in CSV headers: {rows[0]}")
        return None

    dept_idx = find_idx(["department", "dept"], required=True)
    file_no_idx = find_idx(["file no", "file number", "file_no", "file"], required=True)
    item_idx = find_idx(["item name", "item_name", "procurement", "item", "details", "description"], required=True)
    
    unit_cost_idx = find_idx(["unit cost", "unit price", "rate", "cost"])
    qty_idx = find_idx(["quantity", "qty"])
    total_cost_idx = find_idx(["total cost", "budget amount", "total amount", "amount", "total"])
    
    exp_cat_idx = find_idx(["expenditure category", "expenditure type", "expenditure_category"])
    cat_idx = find_idx(["purchase category", "category", "type"])
    course_idx = find_idx(["course code", "course"])
    fy_idx = find_idx(["financial year", "fy", "financial_year"])

    # Extract unique department codes/names from CSV and validate they exist
    csv_depts = set()
    for row in rows[1:]:
        if not row or not any(row):
            continue
        if dept_idx < len(row):
            dept_code = str(row[dept_idx]).strip().lower()
            if dept_code:
                csv_depts.add(dept_code)

    unrecognized_depts = []
    for dept_code in csv_depts:
        dept_res = await db.execute(
            select(Department).where(
                (func.lower(Department.short_code) == dept_code) | 
                (func.lower(Department.name) == dept_code)
            )
        )
        if not dept_res.scalar_one_or_none():
            unrecognized_depts.append(dept_code.upper())

    if unrecognized_depts:
        raise HTTPException(
            status_code=422,
            detail=f"Unrecognized departments in CSV: {', '.join(unrecognized_depts)}"
        )

    # Resolve financial year from form param if provided
    financial_year = None
    if financial_year_id is not None:
        fy_result = await db.execute(select(FinancialYear).where(FinancialYear.id == financial_year_id))
        financial_year = fy_result.scalar_one_or_none()
        if not financial_year:
            raise HTTPException(status_code=400, detail=f"Financial year with ID {financial_year_id} not found")

    # Fallback to active/default if not resolved yet
    if not financial_year:
        now = datetime.utcnow()
        fy_result = await db.execute(
            select(FinancialYear).where(
                and_(FinancialYear.start_date <= now.date(), FinancialYear.end_date >= now.date())
            )
        )
        financial_year = fy_result.scalar_one_or_none()
        if not financial_year:
            fy_result = await db.execute(select(FinancialYear).where(FinancialYear.is_active == True).limit(1))
            financial_year = fy_result.scalar_one_or_none()
        if not financial_year:
            raise HTTPException(status_code=400, detail="No active financial year configured in the system")

    if financial_year.is_closed:
        raise HTTPException(
            status_code=400,
            detail=f"The financial year '{financial_year.label}' is closed. Budgets in closed financial years cannot be imported."
        )

    def clean_float(val_str) -> float:
        if not val_str:
            return 0.0
        cleaned = re.sub(r'[^\d.]', '', val_str)
        return float(cleaned) if cleaned else 0.0

    def clean_int(val_str) -> int:
        if not val_str:
            return 0
        cleaned = re.sub(r'[^\d]', '', val_str)
        return int(cleaned) if cleaned else 0

    success_count = 0
    errors = []
    depts_cache = {}

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        
        try:
            dept_code = str(row[dept_idx]).strip().upper()
            file_no = str(row[file_no_idx]).strip()
            item_name = str(row[item_idx]).strip()

            if not dept_code or not file_no or not item_name:
                errors.append(f"Row {row_num}: Missing required field values (Department, File No, Item Name)")
                continue

            # Determine financial year for this row
            row_fy = financial_year
            if fy_idx is not None and fy_idx < len(row) and row[fy_idx]:
                fy_label = str(row[fy_idx]).strip()
                if fy_label:
                    fy_res = await db.execute(select(FinancialYear).where(FinancialYear.label == fy_label))
                    fy_obj = fy_res.scalar_one_or_none()
                    if fy_obj:
                        row_fy = fy_obj

            if row_fy.is_closed:
                errors.append(f"Row {row_num}: Financial Year '{row_fy.label}' is closed. Budgets in closed financial years cannot be modified.")
                continue

            # Parse amounts and quantities
            unit_cost = 0.0
            quantity = 1
            total_cost = 0.0

            has_unit_cost = unit_cost_idx is not None and unit_cost_idx < len(row) and row[unit_cost_idx]
            has_qty = qty_idx is not None and qty_idx < len(row) and row[qty_idx]
            has_total = total_cost_idx is not None and total_cost_idx < len(row) and row[total_cost_idx]

            if has_unit_cost:
                unit_cost = clean_float(str(row[unit_cost_idx]))
            if has_qty:
                quantity = clean_int(str(row[qty_idx]))
                if quantity <= 0:
                    quantity = 1
            if has_total:
                total_cost = clean_float(str(row[total_cost_idx]))

            # Calculate missing values
            if has_unit_cost and has_qty:
                calculated_total = unit_cost * quantity
                if not has_total or total_cost == 0.0:
                    total_cost = calculated_total
                elif unit_cost == 0.0 and total_cost > 0.0:
                    unit_cost = total_cost / quantity
            elif has_total:
                if not has_unit_cost or unit_cost == 0.0:
                    unit_cost = total_cost
                if not has_qty:
                    quantity = 1

            exp_cat = "CAPEX"
            if exp_cat_idx is not None and exp_cat_idx < len(row) and row[exp_cat_idx]:
                exp_cat = str(row[exp_cat_idx]).strip()

            cat = "equipment"
            if cat_idx is not None and cat_idx < len(row) and row[cat_idx]:
                cat = str(row[cat_idx]).strip()

            course_code = "N/A"
            if course_idx is not None and course_idx < len(row) and row[course_idx]:
                course_code = str(row[course_idx]).strip()

            dept_key = dept_code.lower()
            if dept_key not in depts_cache:
                dept_res = await db.execute(
                    select(Department).where(
                        (func.lower(Department.short_code) == dept_key) | 
                        (func.lower(Department.name) == dept_key)
                    )
                )
                dept = dept_res.scalar_one_or_none()
                if not dept:
                    raise HTTPException(status_code=422, detail=f"Department {dept_code} not found")
                depts_cache[dept_key] = dept.id
            dept_id = depts_cache[dept_key]

            bm_res = await db.execute(
                select(BudgetMaster).where(
                    and_(
                        BudgetMaster.department_id == dept_id,
                        BudgetMaster.file_no == file_no
                    )
                )
            )
            bm = bm_res.scalar_one_or_none()

            if bm:
                bm_fy_res = await db.execute(select(FinancialYear).where(FinancialYear.id == bm.financial_year_id))
                bm_fy = bm_fy_res.scalar_one_or_none()
                if bm_fy and bm_fy.is_closed:
                    errors.append(f"Row {row_num}: Budget file {file_no} already exists in a closed Financial Year '{bm_fy.label}' and cannot be modified.")
                    continue

                bm.total_cost = total_cost
                bm.unit_cost = unit_cost
                bm.item_name = item_name
                bm.quantity = quantity
                bm.financial_year_id = row_fy.id
                if exp_cat_idx is not None:
                    bm.expenditure_category = exp_cat
                if cat_idx is not None:
                    bm.category = cat
                if course_idx is not None:
                    bm.course_code = course_code
            else:
                bm = BudgetMaster(
                    department_id=dept_id,
                    financial_year_id=row_fy.id,
                    expenditure_category=exp_cat,
                    item_name=item_name,
                    category=cat,
                    course_code=course_code,
                    unit_cost=unit_cost,
                    quantity=quantity,
                    total_cost=total_cost,
                    file_no=file_no,
                    is_revision=False,
                )
                db.add(bm)

            success_count += 1
        except Exception as ex:
            errors.append(f"Row {row_num}: Error: {str(ex)}")

    await db.commit()
    return {
        "success": len(errors) == 0,
        "imported": success_count,
        "errors": errors
    }


@router.patch("/users/{user_id}/role")
async def update_user_role(user_id: int, body: dict, db: AsyncSession = Depends(get_db), _=AdminDep):
    role_id = body.get("role_id")
    if not role_id:
        raise HTTPException(status_code=400, detail="role_id is required")
        
    user_res = await db.execute(select(User).where(User.id == user_id))
    u = user_res.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
        
    role_res = await db.execute(select(RoleManager).where(RoleManager.id == role_id))
    role = role_res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=400, detail="Invalid role selected")
        
    u.role_id = role_id
    await db.commit()
    return {"message": "User role updated successfully"}


@router.post("/purchase-requests/{pr_id}/force-advance")
async def force_advance_pr(
    pr_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_roles("admin"))
):
    remarks = body.get("remarks")
    if not remarks or not remarks.strip():
        raise HTTPException(status_code=400, detail="Remarks are mandatory for force-advancing")
        
    result = await db.execute(select(PurchaseRequest).where(PurchaseRequest.id == pr_id))
    pr = result.scalar_one_or_none()
    if not pr:
        raise HTTPException(status_code=404, detail="Purchase request not found")
        
    from app.services.flow_engine import FlowEngineService
    flow_engine = FlowEngineService(db)
    
    try:
        await flow_engine.force_advance(pr, user, remarks)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return {"message": "Purchase request force-advanced successfully"}

