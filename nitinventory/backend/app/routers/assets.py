from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.limiter import limiter
from sqlalchemy import select, or_, func
from datetime import datetime, date
import openpyxl
import weasyprint
import io
import os
import urllib.parse
from app.core.config import settings

from app.core.database import get_db
from app.core.deps import get_current_user, require_roles, require_own_department
from app.models.user import User
from app.models.asset import Asset, AssetMovement, AssetLog
from app.services.asset_service import AssetService
from pydantic import BaseModel
from typing import Optional
from sqlalchemy.orm import selectinload

class PublicAssetView(BaseModel):
    asset_tag: str
    asset_name: str
    location: str
    custodian_name: Optional[str] = None
    department_name: Optional[str] = None

    class Config:
        from_attributes = True

router = APIRouter(prefix="/api/assets", tags=["assets"])


@router.get("/dashboard-stats")
async def get_dashboard_stats(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_own_department())
):
    asset_query = select(Asset)
    if user.role.group_key in ("hod", "faculty"):
        asset_query = asset_query.where(Asset.department_id == user.department_id)
        
    result = await db.execute(asset_query)
    assets = result.scalars().all()
    
    total_assets = len(assets)
    pending_verification = len([a for a in assets if not a.is_verified])
    
    by_category = {}
    for a in assets:
        by_category[a.category] = by_category.get(a.category, 0) + 1
        
    by_condition = {}
    for a in assets:
        by_condition[a.condition] = by_condition.get(a.condition, 0) + 1
        
    by_department = {}
    if user.role.group_key not in ("hod", "faculty"):
        from app.models.user import Department
        dept_result = await db.execute(select(Department))
        depts = {d.id: d.name for d in dept_result.scalars().all()}
        for a in assets:
            dept_name = depts.get(a.department_id, "Unknown Department")
            by_department[dept_name] = by_department.get(dept_name, 0) + 1
            
    recent_query = select(Asset)
    if user.role.group_key in ("hod", "faculty"):
        recent_query = recent_query.where(Asset.department_id == user.department_id)
    recent_query = recent_query.order_by(Asset.created_at.desc()).limit(5)
    
    recent_result = await db.execute(recent_query)
    recent_assets = [
        {
            "id": a.id,
            "asset_tag": a.asset_tag,
            "name": a.name,
            "category": a.category,
            "condition": a.condition,
            "created_at": a.created_at.isoformat()
        }
        for a in recent_result.scalars().all()
    ]
    
    return {
        "total_assets": total_assets,
        "pending_verification": pending_verification,
        "by_category": by_category,
        "by_condition": by_condition,
        "by_department": by_department,
        "recent_assets": recent_assets
    }

@router.get("/")
async def list_assets(
    skip: int = 0,
    limit: int = Query(default=50, le=200),
    search: Optional[str] = None,
    category: Optional[str] = None,
    condition: Optional[str] = None,
    disposal_status: Optional[str] = None,
    fund_source: Optional[str] = None,
    is_verified: Optional[bool] = None,
    department_id: Optional[int] = None,
    year: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_own_department())
):
    base_query = select(Asset)
    if user.role.group_key in ("hod", "faculty"):
        base_query = base_query.where(Asset.department_id == user.department_id)
    elif department_id is not None:
        base_query = base_query.where(Asset.department_id == department_id)
        
    if search:
        search_pattern = f"%{search}%"
        base_query = base_query.where(
            or_(
                Asset.name.ilike(search_pattern),
                Asset.asset_tag.ilike(search_pattern),
                Asset.legacy_asset_tag.ilike(search_pattern),
                Asset.serial_number.ilike(search_pattern),
                Asset.custodian.ilike(search_pattern),
                Asset.building.ilike(search_pattern),
                Asset.room.ilike(search_pattern),
            )
        )
        
    if category:
        base_query = base_query.where(Asset.category == category)
    if condition:
        base_query = base_query.where(Asset.condition == condition)
    if disposal_status:
        base_query = base_query.where(Asset.disposal_status == disposal_status)
    if fund_source:
        base_query = base_query.where(Asset.fund_source == fund_source)
    if is_verified is not None:
        base_query = base_query.where(Asset.is_verified == is_verified)
    if year is not None:
        year_suffix = f"-{str(year)[-2:]}-"
        base_query = base_query.where(Asset.asset_tag.like(f"%{year_suffix}%"))
    
    count_query = select(func.count()).select_from(base_query.subquery())
    total = await db.scalar(count_query) or 0

    query = base_query.order_by(Asset.created_at.desc(), Asset.id.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    assets = result.scalars().all()
    
    items = [
        {"id": a.id, "asset_tag": a.asset_tag, "legacy_asset_tag": a.legacy_asset_tag, "fund_source": a.fund_source, "name": a.name, "category": a.category,
         "condition": a.condition, "disposal_status": a.disposal_status,
         "building": a.building, "room": a.room, "qr_code_url": a.qr_code_url,
         "delivery_item_id": a.delivery_item_id, "department_id": a.department_id,
         "serial_number": a.serial_number, "custodian": a.custodian,
         "remarks": a.remarks, "is_verified": a.is_verified, "asset_source": a.asset_source,
         "purchase_date": a.purchase_date.isoformat() if a.purchase_date else None,
         "unit_cost": a.unit_cost,
         "quantity": a.quantity,
         "supplier_name": a.supplier_name,
         "supplier_address": a.supplier_address,
         "bill_number": a.bill_number,
         "bill_date": a.bill_date.isoformat() if a.bill_date else None,
         "delivery_date": a.delivery_date.isoformat() if a.delivery_date else None,
         "stock_register_volume": a.stock_register_volume,
         "stock_register_page": a.stock_register_page}
        for a in assets
    ]
    return {"items": items, "total": total}


@router.post("/")
async def register_asset(body: dict, db: AsyncSession = Depends(get_db), user: User = Depends(require_roles("hod", "admin"))):
    svc = AssetService(db)
    asset = await svc.register_asset(body, user)
    await db.commit()
    return {
        "message": "Asset manually registered successfully",
        "id": asset.id,
        "asset_tag": asset.asset_tag,
        "legacy_asset_tag": asset.legacy_asset_tag,
        "fund_source": asset.fund_source
    }


@router.get("/qr/{asset_tag}", response_model=PublicAssetView)
async def public_asset_profile(asset_tag: str, db: AsyncSession = Depends(get_db)):
    """Public route — no auth. Accessible via QR scan."""
    result = await db.execute(
        select(Asset)
        .options(selectinload(Asset.department))
        .where(Asset.asset_tag == asset_tag)
    )
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {
        "asset_tag": asset.asset_tag,
        "asset_name": asset.name,
        "location": f"{asset.building or ''} {asset.room or ''}".strip(),
        "custodian_name": asset.custodian,
        "department_name": asset.department.name if asset.department else None,
    }


@router.get("/{asset_id}")
async def get_asset(asset_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    result = await db.execute(
        select(Asset)
        .options(
            selectinload(Asset.movements),
            selectinload(Asset.logs).selectinload(AssetLog.performed_by)
        )
        .where(Asset.id == asset_id)
    )
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if user.role.group_key in ("hod", "faculty") and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
    return {
        "id": asset.id, "asset_tag": asset.asset_tag, "legacy_asset_tag": asset.legacy_asset_tag, "fund_source": asset.fund_source, "name": asset.name,
        "category": asset.category, "condition": asset.condition,
        "disposal_status": asset.disposal_status, "qr_code_url": asset.qr_code_url,
        "building": asset.building, "room": asset.room, "custodian": asset.custodian,
        "serial_number": asset.serial_number, "unit_cost": asset.unit_cost,
        "remarks": asset.remarks, "is_verified": asset.is_verified,
        "verified_at": asset.verified_at.isoformat() if asset.verified_at else None,
        "asset_source": asset.asset_source,
        "purchase_date": asset.purchase_date.isoformat() if asset.purchase_date else None,
        "warranty_expiry": asset.warranty_expiry.isoformat() if asset.warranty_expiry else None,
        "quantity": asset.quantity,
        "supplier_name": asset.supplier_name,
        "supplier_address": asset.supplier_address,
        "bill_number": asset.bill_number,
        "bill_date": asset.bill_date.isoformat() if asset.bill_date else None,
        "delivery_date": asset.delivery_date.isoformat() if asset.delivery_date else None,
        "stock_register_volume": asset.stock_register_volume,
        "stock_register_page": asset.stock_register_page,
        "movements": [{"from_room": m.from_room, "to_room": m.to_room, "moved_at": m.moved_at.isoformat(), "reason": m.reason} for m in asset.movements],
        "logs": [
            {
                "action": l.action,
                "performed_at": l.performed_at.isoformat(),
                "old_value": l.old_value,
                "new_value": l.new_value,
                "performed_by_name": l.performed_by.name if l.performed_by else f"User {l.performed_by_id}"
            }
            for l in asset.logs
        ],
    }

@router.put("/{asset_id}")
async def update_asset(asset_id: int, body: dict, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
        
    if user.role.group_key in ("hod", "faculty") and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
        
    old_values = {
        "name": asset.name,
        "category": asset.category,
        "building": asset.building,
        "room": asset.room,
        "custodian": asset.custodian,
        "serial_number": asset.serial_number,
        "legacy_asset_tag": asset.legacy_asset_tag,
        "unit_cost": asset.unit_cost,
        "remarks": asset.remarks,
        "fund_source": asset.fund_source,
        "asset_source": asset.asset_source,
        "quantity": asset.quantity,
        "supplier_name": asset.supplier_name,
        "supplier_address": asset.supplier_address,
        "bill_number": asset.bill_number,
        "bill_date": asset.bill_date.isoformat() if asset.bill_date else None,
        "delivery_date": asset.delivery_date.isoformat() if asset.delivery_date else None,
        "stock_register_volume": asset.stock_register_volume,
        "stock_register_page": asset.stock_register_page,
    }
    
    if "name" in body:
        asset.name = body["name"]
    if "category" in body:
        asset.category = body["category"]
    if "building" in body:
        asset.building = body["building"]
    if "room" in body:
        asset.room = body["room"]
    if "custodian" in body:
        asset.custodian = body["custodian"]
    if "serial_number" in body:
        asset.serial_number = body["serial_number"]
    if "legacy_asset_tag" in body:
        asset.legacy_asset_tag = body["legacy_asset_tag"]
    if "unit_cost" in body:
        asset.unit_cost = float(body["unit_cost"]) if body["unit_cost"] is not None else None
    if "remarks" in body:
        asset.remarks = body["remarks"]
    if "fund_source" in body:
        asset.fund_source = body["fund_source"]
    if "asset_source" in body:
        asset.asset_source = body["asset_source"]
    if "quantity" in body:
        asset.quantity = int(body["quantity"]) if body["quantity"] is not None else 1
    if "supplier_name" in body:
        asset.supplier_name = body["supplier_name"]
    if "supplier_address" in body:
        asset.supplier_address = body["supplier_address"]
    if "bill_number" in body:
        asset.bill_number = body["bill_number"]
    if "stock_register_volume" in body:
        asset.stock_register_volume = body["stock_register_volume"]
    if "stock_register_page" in body:
        asset.stock_register_page = body["stock_register_page"]
        
    if "purchase_date" in body:
        if body["purchase_date"]:
            asset.purchase_date = datetime.strptime(body["purchase_date"], "%Y-%m-%d").date()
        else:
            asset.purchase_date = None
            
    if "warranty_expiry" in body:
        if body["warranty_expiry"]:
            asset.warranty_expiry = datetime.strptime(body["warranty_expiry"], "%Y-%m-%d").date()
        else:
            asset.warranty_expiry = None

    if "bill_date" in body:
        if body["bill_date"]:
            asset.bill_date = datetime.strptime(body["bill_date"], "%Y-%m-%d").date()
        else:
            asset.bill_date = None

    if "delivery_date" in body:
        if body["delivery_date"]:
            asset.delivery_date = datetime.strptime(body["delivery_date"], "%Y-%m-%d").date()
        else:
            asset.delivery_date = None
        
    new_values = {
        "name": asset.name,
        "category": asset.category,
        "building": asset.building,
        "room": asset.room,
        "custodian": asset.custodian,
        "serial_number": asset.serial_number,
        "legacy_asset_tag": asset.legacy_asset_tag,
        "unit_cost": asset.unit_cost,
        "remarks": asset.remarks,
        "fund_source": asset.fund_source,
        "asset_source": asset.asset_source,
        "quantity": asset.quantity,
        "supplier_name": asset.supplier_name,
        "supplier_address": asset.supplier_address,
        "bill_number": asset.bill_number,
        "bill_date": asset.bill_date.isoformat() if asset.bill_date else None,
        "delivery_date": asset.delivery_date.isoformat() if asset.delivery_date else None,
        "stock_register_volume": asset.stock_register_volume,
        "stock_register_page": asset.stock_register_page,
    }
    
    log = AssetLog(
        asset_id=asset.id,
        action="asset_updated",
        performed_by_id=user.id,
        old_value=old_values,
        new_value=new_values,
        performed_at=datetime.utcnow(),
    )
    db.add(log)
    await db.commit()
    return {"message": "Asset updated successfully"}

@router.post("/{asset_id}/verify")
async def verify_asset(asset_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if user.role.group_key in ("hod", "faculty") and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
        
    asset.is_verified = True
    asset.verified_at = datetime.utcnow()
    
    log = AssetLog(
        asset_id=asset.id,
        action="asset_verified",
        performed_by_id=user.id,
        old_value={"is_verified": False},
        new_value={"is_verified": True, "verified_at": asset.verified_at.isoformat()},
        performed_at=datetime.utcnow(),
    )
    db.add(log)
    await db.commit()
    return {"message": "Asset physically verified successfully", "is_verified": True}


@router.patch("/{asset_id}/condition")
async def update_condition(asset_id: int, body: dict, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if user.role.group_key in ("hod", "faculty") and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
    svc = AssetService(db)
    asset = await svc.update_condition(asset_id, body["condition"], user)
    await db.commit()
    return {"message": "Condition updated", "condition": asset.condition}


@router.post("/{asset_id}/move")
async def move_asset(asset_id: int, body: dict, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if user.role.group_key in ("hod", "faculty") and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
    svc = AssetService(db)
    await svc.move_asset(asset_id, body["to_building"], body["to_room"], user, body.get("reason"))
    await db.commit()
    return {"message": "Asset movement recorded"}


@router.post("/{asset_id}/flag-disposal")
async def flag_disposal(asset_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    if user.role.group_key not in ("hod", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if user.role.group_key == "hod" and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
    svc = AssetService(db)
    asset = await svc.flag_disposal(asset_id, user)
    await db.commit()
    return {"message": "Asset flagged for disposal", "disposal_status": asset.disposal_status}


@router.post("/{asset_id}/confirm-disposal")
async def confirm_disposal(asset_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    if user.role.group_key != "admin":
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    svc = AssetService(db)
    asset = await svc.confirm_disposal(asset_id, user)
    await db.commit()
    return {"message": "Disposal confirmed", "disposal_status": asset.disposal_status}


@router.delete("/{asset_id}")
async def delete_asset(asset_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(require_own_department())):
    if user.role.group_key not in ("hod", "admin"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if user.role.group_key == "hod" and asset.department_id != user.department_id:
        raise HTTPException(status_code=403, detail="Access denied")
    svc = AssetService(db)
    await svc.delete_asset(asset_id, user)
    await db.commit()
    return {"message": "Asset deleted successfully"}


@router.post("/import")
@limiter.limit("10/minute")
async def import_assets(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_roles())
):
    contents = await file.read()
    file_content = contents.decode("utf-8")
    svc = AssetService(db)
    result = await svc.import_assets_csv(file_content, user)
    await db.commit()
    return result


@router.get("/export/excel")
async def export_assets_excel(
    search: Optional[str] = None,
    category: Optional[str] = None,
    condition: Optional[str] = None,
    disposal_status: Optional[str] = None,
    fund_source: Optional[str] = None,
    is_verified: Optional[bool] = None,
    department_id: Optional[int] = None,
    year: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_own_department())
):
    base_query = select(Asset)
    if user.role.group_key in ("hod", "faculty"):
        base_query = base_query.where(Asset.department_id == user.department_id)
    elif department_id is not None:
        base_query = base_query.where(Asset.department_id == department_id)
        
    if search:
        search_pattern = f"%{search}%"
        base_query = base_query.where(
            or_(
                Asset.name.ilike(search_pattern),
                Asset.asset_tag.ilike(search_pattern),
                Asset.legacy_asset_tag.ilike(search_pattern),
                Asset.serial_number.ilike(search_pattern),
                Asset.custodian.ilike(search_pattern),
                Asset.building.ilike(search_pattern),
                Asset.room.ilike(search_pattern),
            )
        )
        
    if category:
        base_query = base_query.where(Asset.category == category)
    if condition:
        base_query = base_query.where(Asset.condition == condition)
    if disposal_status:
        base_query = base_query.where(Asset.disposal_status == disposal_status)
    if fund_source:
        base_query = base_query.where(Asset.fund_source == fund_source)
    if is_verified is not None:
        base_query = base_query.where(Asset.is_verified == is_verified)
    if year is not None:
        year_suffix = f"-{str(year)[-2:]}-"
        base_query = base_query.where(Asset.asset_tag.like(f"%{year_suffix}%"))
        
    base_query = base_query.options(selectinload(Asset.department))
    query = base_query.order_by(Asset.created_at.desc(), Asset.id.desc())
    result = await db.execute(query)
    assets = result.scalars().all()
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Physical Asset Register"

    title_font = Font(name="Calibri", size=16, bold=True, color="1A3A6B")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )

    ws.append(["NATIONAL INSTITUTE OF TECHNOLOGY, TIRUCHIRAPPALLI"])
    ws.append(["DEPARTMENTAL PHYSICAL ASSET REGISTER"])
    ws.append([f"Report Generated: {datetime.now().strftime('%d/%m/%Y')} | Generated By: {user.name}"])
    ws.append([]) # Blank row

    ws.merge_cells("A1:X1")
    ws.merge_cells("A2:X2")
    ws.merge_cells("A3:X3")
    
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(name="Calibri", size=13, bold=True, color="333333")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = [
        "SL.NO.", "Asset Tag", "Existing/Legacy Tag", "Asset Name", "Category", 
        "Department", "Funding Source", "Unit Cost (Rs.)", "Quantity", 
        "Building Location", "Room Location", "Custodian / In-Charge", 
        "Serial Number", "Supplier Name", "Supplier Address", "Bill Number", 
        "Bill Date", "Delivery Date", "Stock Register Volume", "Stock Register Page", 
        "Condition", "Purchase Date", "Warranty Expiry", "Remarks"
    ]
    ws.append(headers)
    
    header_row_idx = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[header_row_idx].height = 28

    for idx, a in enumerate(assets, 1):
        row_idx = header_row_idx + idx
        row_data = [
            idx,
            a.asset_tag,
            a.legacy_asset_tag or "",
            a.name,
            a.category.replace("_", " ").title() if a.category else "",
            a.department.name if a.department else "",
            a.fund_source.replace("_", " ").title() if a.fund_source else "",
            a.unit_cost if a.unit_cost is not None else "",
            a.quantity,
            a.building or "",
            a.room or "",
            a.custodian or "",
            a.serial_number or "",
            a.supplier_name or "",
            a.supplier_address or "",
            a.bill_number or "",
            a.bill_date.strftime("%Y-%m-%d") if a.bill_date else "",
            a.delivery_date.strftime("%Y-%m-%d") if a.delivery_date else "",
            a.stock_register_volume or "",
            a.stock_register_page or "",
            a.condition.replace("_", " ").upper() if a.condition else "WORKING",
            a.purchase_date.strftime("%Y-%m-%d") if a.purchase_date else "",
            a.warranty_expiry.strftime("%Y-%m-%d") if a.warranty_expiry else "",
            a.remarks or ""
        ]
        ws.append(row_data)
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 8, 9, 16, 17, 18, 19, 20, 21, 22, 23):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
            if idx % 2 == 0:
                cell.fill = zebra_fill

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[4:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"asset_register_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf")
async def export_assets_pdf(
    search: Optional[str] = None,
    category: Optional[str] = None,
    condition: Optional[str] = None,
    disposal_status: Optional[str] = None,
    fund_source: Optional[str] = None,
    is_verified: Optional[bool] = None,
    department_id: Optional[int] = None,
    year: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_own_department())
):
    base_query = select(Asset)
    if user.role.group_key in ("hod", "faculty"):
        base_query = base_query.where(Asset.department_id == user.department_id)
    elif department_id is not None:
        base_query = base_query.where(Asset.department_id == department_id)
        
    if search:
        search_pattern = f"%{search}%"
        base_query = base_query.where(
            or_(
                Asset.name.ilike(search_pattern),
                Asset.asset_tag.ilike(search_pattern),
                Asset.legacy_asset_tag.ilike(search_pattern),
                Asset.serial_number.ilike(search_pattern),
                Asset.custodian.ilike(search_pattern),
                Asset.building.ilike(search_pattern),
                Asset.room.ilike(search_pattern),
            )
        )
        
    if category:
        base_query = base_query.where(Asset.category == category)
    if condition:
        base_query = base_query.where(Asset.condition == condition)
    if disposal_status:
        base_query = base_query.where(Asset.disposal_status == disposal_status)
    if fund_source:
        base_query = base_query.where(Asset.fund_source == fund_source)
    if is_verified is not None:
        base_query = base_query.where(Asset.is_verified == is_verified)
    if year is not None:
        year_suffix = f"-{str(year)[-2:]}-"
        base_query = base_query.where(Asset.asset_tag.like(f"%{year_suffix}%"))
        
    base_query = base_query.options(selectinload(Asset.department))
    query = base_query.order_by(Asset.created_at.desc(), Asset.id.desc())
    result = await db.execute(query)
    assets = result.scalars().all()

    dept_name = "Central Administration"
    if user.role.group_key in ("hod", "faculty") and user.department:
        dept_name = user.department.name
    elif department_id is not None:
        from app.models.user import Department
        d_res = await db.execute(select(Department).where(Department.id == department_id))
        d_obj = d_res.scalar_one_or_none()
        if d_obj:
            dept_name = d_obj.name

    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "NITLOGO.png")
    logo_url = f"file://{urllib.parse.quote(logo_path, safe='/')}" if os.path.exists(logo_path) else ""

    rows_html = ""
    for i, a in enumerate(assets, 1):
        rows_html += f"""
        <tr>
            <td class="text-center">{i}</td>
            <td class="font-mono">{a.asset_tag}</td>
            <td class="font-mono">{a.legacy_asset_tag or '—'}</td>
            <td><b>{a.name}</b></td>
            <td>{a.category.replace('_', ' ').title() if a.category else '—'}</td>
            <td>{a.department.name if a.department else '—'}</td>
            <td>{a.fund_source.replace('_', ' ').title() if a.fund_source else '—'}</td>
            <td class="text-right">Rs. {f"{int(a.unit_cost):,}" if a.unit_cost is not None else '—'}</td>
            <td class="text-center">{a.quantity}</td>
            <td>{a.building or '—'}</td>
            <td>{a.room or '—'}</td>
            <td>{a.custodian or '—'}</td>
            <td class="font-mono">{a.serial_number or '—'}</td>
            <td>{a.supplier_name or '—'}</td>
            <td>{a.supplier_address or '—'}</td>
            <td>{a.bill_number or '—'}</td>
            <td class="text-center">{a.bill_date.strftime('%d/%m/%Y') if a.bill_date else '—'}</td>
            <td class="text-center">{a.delivery_date.strftime('%d/%m/%Y') if a.delivery_date else '—'}</td>
            <td class="text-center">{a.stock_register_volume or '—'}</td>
            <td class="text-center">{a.stock_register_page or '—'}</td>
            <td class="text-center">{a.condition.upper() if a.condition else 'WORKING'}</td>
            <td class="text-center">{a.purchase_date.strftime('%d/%m/%Y') if a.purchase_date else '—'}</td>
            <td class="text-center">{a.warranty_expiry.strftime('%d/%m/%Y') if a.warranty_expiry else '—'}</td>
            <td>{a.remarks or '—'}</td>
        </tr>
        """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        @page {{ size: A3 landscape; margin: 0.5cm; }}
        body {{ font-family: Arial, sans-serif; font-size: 6pt; color: #333; }}
        .header-table {{ width: 100%; border-collapse: collapse; margin-bottom: 10px; }}
        .logo {{ height: 40px; }}
        .title {{ font-size: 10pt; font-weight: bold; color: #1a3a6b; text-transform: uppercase; }}
        .register-table {{ width: 100%; border-collapse: collapse; }}
        .register-table th {{ background-color: #1a3a6b; color: white; padding: 3px; border: 0.5px solid #ccc; }}
        .register-table td {{ border: 0.5px solid #ccc; padding: 3px; }}
        .text-center {{ text-align: center; }}
        .text-right {{ text-align: right; }}
        .font-mono {{ font-family: monospace; }}
    </style>
    </head>
    <body>
        <table class="header-table">
            <tr>
                <td style="width: 50px;"><img class="logo" src="{logo_url}" /></td>
                <td class="text-center"><div class="title">Physical Asset Register</div><div>{dept_name} | Generated: {datetime.now().strftime('%d/%m/%Y')}</div></td>
            </tr>
        </table>
        <table class="register-table">
            <thead>
                <tr>
                    <th>SL.NO.</th><th>Tag</th><th>Legacy</th><th>Name</th><th>Category</th><th>Dept</th><th>Source</th>
                    <th>Cost</th><th>Qty</th><th>Bldg</th><th>Room</th><th>Custodian</th>
                    <th>Serial</th><th>Supplier</th><th>Address</th><th>Bill No</th><th>Bill Dt</th><th>Del Dt</th>
                    <th>Vol</th><th>Pg</th><th>Cond</th><th>Pur Dt</th><th>Warranty</th><th>Remarks</th>
                </tr>
            </thead>
            <tbody>{rows_html}</tbody>
        </table>
    </body>
    </html>
    """
    
    try:
        pdf_bytes = weasyprint.HTML(string=html_content).write_pdf()
        filename = f"asset_register_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception:
        return HTMLResponse(content=html_content)
